# scripts/datasets3.py
import os
import csv
import random
import numpy as np
import cv2

import torch
import torch.nn.functional as F
import torchvision.transforms as T
import torchaudio
from torch.utils.data import Dataset
from openpyxl import load_workbook

import logging
logging.getLogger('moviepy').setLevel(logging.CRITICAL)
logging.getLogger('moviepy.video.io.ffmpeg_reader').setLevel(logging.CRITICAL)
os.environ['FFMPEG_LOGLEVEL'] = 'quiet'


class LIVESJTUAVQADataset(Dataset):
    """
    LIVE-SJTU_AVQA dataset loader (Distorted/*.yuv + Distorted/*.wav) using MOS.xlsx

    Returns:
      (video_frames, waveform, audio_reliability_tensor, normalized_mos, sample_name)

    Notes:
    - YUV is assumed YUV420p (I420)
    - resolution is read from MOS.xlsx sheet 'resolution'
    - frame count per sample read from MOS.xlsx sheet 'frameNum'
    - MOS is assumed in [0,100] -> normalized to [0,1]
    - reliability is read from CSV (floor to >= 0.1), fallback 0.5
    """

    def __init__(
        self,
        video_dir="datasets/LIVE-SJTU_AVQA/Distorted",
        mos_xlsx_path="datasets/LIVE-SJTU_AVQA/MOS.xlsx",
        scoreq_reliability_csv_path=None,
        train=False
    ):
        self.video_dir = video_dir
        self.train = train

        self.num_frames = 8
        self.out_size = 224

        self.audio_sr = 16000
        self.max_audio_len = 672000  # 42 sec @ 16kHz

        # -----------------------------
        # Load reliability CSV
        # -----------------------------
        self.reliability_map = {}
        if scoreq_reliability_csv_path and os.path.isfile(scoreq_reliability_csv_path):
            with open(scoreq_reliability_csv_path, "r", newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    fn = (row.get("filename") or "").strip()
                    if not fn:
                        continue
                    key = os.path.splitext(os.path.basename(fn))[0].strip()
                    try:
                        rel = float(row.get("scoreq_reliability", 0.5))
                    except Exception:
                        rel = 0.5
                    if not np.isfinite(rel):
                        rel = 0.5
                    rel = float(np.clip(rel, 0.1, 1.0))
                    self.reliability_map[key] = rel

            print(f"✅ Loaded ScoreQ reliability for {len(self.reliability_map)} samples from {scoreq_reliability_csv_path}")
            print(f"   Example entry: {list(self.reliability_map.items())[0] if self.reliability_map else 'N/A'}")
        else:
            print("⚠️ No reliability CSV found. Using default reliability=0.5 for all samples.")

        # -----------------------------
        # Load MOS.xlsx
        # -----------------------------
        wb = load_workbook(mos_xlsx_path, data_only=True)

        dis_ws = wb["disNames"]
        mos_ws = wb["MOS"]
        fn_ws = wb["frameNum"]
        res_ws = wb["resolution"]

        # resolution sheet: first row contains H,W
        H = res_ws["A1"].value
        W = res_ws["B1"].value
        self.height = int(H)
        self.width = int(W)
        print(f"✅ LIVE-SJTU resolution from MOS.xlsx: H={self.height}, W={self.width}")

        # -----------------------------
        # Build samples list
        # -----------------------------
        self.samples = []
        missing_files = 0
        missing_rel = 0

        # NOTE: rows are aligned by index in excel
        # Each disNames row: [dist_yuv, dist_wav]
        for i in range(1, dis_ws.max_row + 1):
            yuv_name = dis_ws.cell(row=i, column=1).value
            wav_name = dis_ws.cell(row=i, column=2).value
            mos_val = mos_ws.cell(row=i, column=1).value
            frame_num = fn_ws.cell(row=i, column=1).value

            if yuv_name is None or wav_name is None or mos_val is None or frame_num is None:
                continue

            yuv_name = str(yuv_name).strip()
            wav_name = str(wav_name).strip()

            try:
                mos_val = float(mos_val)
            except Exception:
                continue
            if not np.isfinite(mos_val):
                continue

            try:
                frame_num = int(float(frame_num))
            except Exception:
                frame_num = 0

            yuv_path = os.path.join(self.video_dir, yuv_name)
            wav_path = os.path.join(self.video_dir, wav_name)

            if (not os.path.isfile(yuv_path)) or (not os.path.isfile(wav_path)):
                missing_files += 1
                continue

            wav_key = os.path.splitext(os.path.basename(wav_name))[0].strip()
            if scoreq_reliability_csv_path and (wav_key not in self.reliability_map):
                missing_rel += 1
                continue

            name = f"{os.path.splitext(yuv_name)[0]}__{wav_key}"

            self.samples.append({
                "yuv": yuv_path,
                "wav": wav_path,
                "mos": mos_val,
                "frames": frame_num,
                "wav_key": wav_key,
                "name": name
            })

        print(f"✅ Built LIVE-SJTU samples: {len(self.samples)}")
        if missing_files > 0:
            print(f"⚠️ Skipped {missing_files} rows due to missing yuv/wav files")
        if missing_rel > 0:
            print(f"⚠️ Skipped {missing_rel} rows due to missing reliability entries")

        if len(self.samples) == 0:
            raise ValueError("❌ No valid LIVE-SJTU samples found. Check MOS.xlsx alignment + paths.")

        # -----------------------------
        # Augmentation (same style as datasets.py)
        # -----------------------------
        self.video_transform = T.Compose([
            T.ToPILImage(),
            T.RandomResizedCrop(224, scale=(0.8, 1.0)),
            T.RandomRotation(degrees=15),
            T.RandomHorizontalFlip(),
            T.ToTensor(),
        ])

        self.audio_transforms = [
            lambda x: x + 0.005 * torch.randn_like(x) if random.random() < 0.5 else x,
            lambda x: x * random.uniform(0.8, 1.2) if random.random() < 0.5 else x,
        ]

    def __len__(self):
        return len(self.samples)

    def _load_yuv_frames(self, yuv_path, total_frames):
        if total_frames <= 0:
            return torch.zeros(self.num_frames, 3, self.out_size, self.out_size)

        frame_size = int(self.width * self.height * 3 // 2)  # YUV420p
        indices = np.linspace(0, total_frames - 1, self.num_frames).astype(int).tolist()

        frames = []
        try:
            with open(yuv_path, "rb") as f:
                for fi in indices:
                    f.seek(fi * frame_size, os.SEEK_SET)
                    raw = f.read(frame_size)
                    if raw is None or len(raw) != frame_size:
                        frames.append(None)
                        continue
                    yuv = np.frombuffer(raw, dtype=np.uint8).reshape((self.height * 3 // 2, self.width))
                    rgb = cv2.cvtColor(yuv, cv2.COLOR_YUV2RGB_I420)
                    rgb = cv2.resize(rgb, (self.out_size, self.out_size), interpolation=cv2.INTER_AREA)
                    frames.append(rgb)
        except Exception:
            return torch.zeros(self.num_frames, 3, self.out_size, self.out_size)

        if all(fr is None for fr in frames):
            return torch.zeros(self.num_frames, 3, self.out_size, self.out_size)

        first = next(fr for fr in frames if fr is not None)
        prev = first
        fixed = []
        for fr in frames:
            if fr is None:
                fr = prev
            else:
                prev = fr
            fixed.append(fr)

        arr = np.stack(fixed, axis=0).astype(np.float32) / 255.0  # [T,H,W,3]
        vid = torch.from_numpy(arr).permute(0, 3, 1, 2)  # [T,3,H,W]
        return vid

    def _load_audio(self, wav_path):
        try:
            waveform, sr = torchaudio.load(wav_path)  # [C,T]
            if waveform.ndim == 2 and waveform.size(0) > 1:
                waveform = waveform.mean(dim=0)
            else:
                waveform = waveform.squeeze(0)
            waveform = waveform.float()

            if sr != self.audio_sr:
                resampler = torchaudio.transforms.Resample(sr, self.audio_sr)
                waveform = resampler(waveform)

        except Exception as e:
            print(f"⚠️  Error loading wav {wav_path}: {e}")
            waveform = torch.zeros(self.max_audio_len, dtype=torch.float32)

        if torch.isnan(waveform).any() or torch.isinf(waveform).any():
            waveform = torch.zeros(self.max_audio_len, dtype=torch.float32)

        cur_len = waveform.shape[0]
        if cur_len > self.max_audio_len:
            waveform = waveform[:self.max_audio_len]
        elif cur_len < self.max_audio_len:
            waveform = F.pad(waveform, (0, self.max_audio_len - cur_len))

        if not torch.isfinite(waveform).all():
            waveform = torch.zeros_like(waveform)

        if self.train:
            for aug in self.audio_transforms:
                try:
                    before = waveform.clone()
                    waveform = aug(waveform)
                    if torch.isnan(waveform).any() or torch.isinf(waveform).any():
                        waveform = before
                except Exception:
                    pass

        return waveform

    def __getitem__(self, idx):
        s = self.samples[idx]

        # MOS: [0..100] -> [0,1]
        mos = float(np.clip(s["mos"] / 100.0, 0.0, 1.0))

        # reliability: clip >= 0.1
        rel = self.reliability_map.get(s["wav_key"], 0.5)
        if not np.isfinite(rel):
            rel = 0.5
        rel = float(np.clip(rel, 0.1, 1.0))

        video_frames = self._load_yuv_frames(s["yuv"], s["frames"])
        if torch.isnan(video_frames).any() or torch.isinf(video_frames).any():
            video_frames = torch.zeros(self.num_frames, 3, self.out_size, self.out_size)

        if self.train:
            aug_frames = []
            for f in video_frames:
                f = self.video_transform((f.clamp(0, 1) * 255).byte())
                aug_frames.append(f)
            video_frames = torch.stack(aug_frames, dim=0)

        waveform = self._load_audio(s["wav"])

        # 🔀 conditional return
        if self.reliability_map:
            # for models that use reliability
            return (
                video_frames,
                waveform,
                torch.tensor([rel], dtype=torch.float32),
                mos,
                s["name"]
            )
        else:
            # ✅ baseline: no reliability at all
            return (
                video_frames,
                waveform,
                mos,
                s["name"]
            )